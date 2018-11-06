from copy import copy

from openpyxl.cell import Cell
from openpyxl.worksheet import Worksheet

import openpyxl

from wysiwygtemplate.dictcontext import DictExcelTemplateContext
from wysiwygtemplate.looptemplate import ExcelArchetectureTemplate
from wysiwygtemplate.pyevaluator import PyEvaluator, EmbeddedPyEvaluator
from wysiwygtemplate.replacetemplate import ExcelRelpaceTemplate
from wysiwygtemplate.templatebase import ExcelProcessor

def borderfix(workbook:openpyxl.Workbook):
    for sheet in workbook.worksheets:
        for merged_cell in sheet.merged_cells:
            border = copy(sheet.cell(merged_cell.min_row, merged_cell.min_col).border)
            for row in range(merged_cell.min_row, merged_cell.max_row+1):
                for col in range(merged_cell.min_col, merged_cell.max_col+1):
                    sheet.cell(row, col).border = copy(border)
            pass

def copyCellFormat(cellsrc:Cell, sheetSrc:Worksheet, celldest:Cell, sheetDes:Worksheet):
    celldest.fill = copy(cellsrc.fill)
    celldest.font = copy(cellsrc.font)
    celldest.border = copy(cellsrc.border)
    celldest.alignment = copy(cellsrc.alignment)
    celldest.number_format = copy(cellsrc.number_format)
    celldest.protection = copy(cellsrc.protection)

    for merged_cell in sheetSrc.merged_cells:
        if merged_cell.min_col==cellsrc.col_idx and merged_cell.min_row==cellsrc.row:
            sheetDes.merge_cells(start_row= celldest.row, end_row= celldest.row,
                                 start_column= merged_cell.min_col, end_column= merged_cell.max_col)
            break

class OpenpyXlExcelProcessor(ExcelProcessor):
    def __init__(self, xlsfilepath, sheetnameorindex) -> None:
        super().__init__(xlsfilepath)
        self.workbook = openpyxl.load_workbook(xlsfilepath)
        if isinstance(sheetnameorindex, int):
            self.worksheet = self.workbook.worksheets[sheetnameorindex]
        elif isinstance(sheetnameorindex, str):
            self.worksheet = self.workbook.get_sheet_by_name(sheetnameorindex)

    def bound(self) -> tuple:
        s:Worksheet = self.worksheet
        return s.min_row, s.max_row, s.min_column, s.max_column

    def insertTemplate(self, template, context, row, col):
        super().insertTemplate(template, context, row, col)

    def readCellValue(self, row, col):
        return self.worksheet.cell(row, col).value

    def writeCellValue(self, row, col, val):
        self.worksheet.cell(row, col).value = val

    def insertRowBefore(self, rowdata, row):
        sheet: Worksheet = self.worksheet
        sheet.insert_rows(row, amount= len(rowdata))
        for i in range(len(rowdata)):
            for c in range(sheet.min_column,sheet.max_column+1):
                cell = sheet.cell(row + i, c)
                copyCellFormat(sheet.cell(row+len(rowdata), c), sheet, cell, sheet)
        for i in range(len(rowdata)):
            rowitem = rowdata[i]
            for c in range(sheet.min_column,sheet.max_column+1):
                if c-sheet.min_column>=len(rowitem): continue
                sheet.cell(row+i, c, rowitem[c-sheet.min_column])

        for i in range(len(rowdata)):
            sheet.row_dimensions[row + len(rowdata) + i].height = sheet.row_dimensions[row + i].height if sheet.row_dimensions[row + i].height is not None else 15
        for i in range(1, len(rowdata)):
            sheet.row_dimensions[row + i].height = sheet.row_dimensions[row].height if sheet.row_dimensions[row].height is not None else 15

        mergedmap = {}
        for r in sheet.merged_cells.ranges:
            if r.min_row< row:
                if r.max_row>= row:
                    mergedmap[(r.min_row, r.max_row, r.min_col, r.max_col)] = (r.min_row, r.max_row+len(rowdata), r.min_col, r.max_col)
            else:
                mergedmap[(r.min_row, r.max_row, r.min_col, r.max_col)] = (r.min_row+len(rowdata), r.max_row + len(rowdata), r.min_col, r.max_col)
        for min_row, max_row, min_col, max_col in mergedmap.keys():
            sheet.unmerge_cells(start_row= min_row, start_column=min_col, end_row=max_row, end_column= max_col)

        for min_row, max_row, min_col, max_col in mergedmap.values():
            sheet.merge_cells(start_row= min_row, start_column=min_col, end_row=max_row, end_column= max_col)
        pass



    def insertRowAfter(self, rowdata, row):
        pass

    def insertColumnBefore(self, coldata, row):
        super().insertColumnBefore(coldata, row)

    def insertColumnAfter(self, coldata, row):
        super().insertColumnAfter(coldata, row)

    def deleteRow(self, row):
        sheet: Worksheet = self.worksheet
        sheet.delete_rows(row- sheet.min_row, 1)
        removelist = []
        map = {}
        for r in sheet.merged_cells.ranges:
            if r.min_row<= row:
                if r.min_row==r.max_row==row: removelist.append(r)
                elif r.max_row>= row:
                    map[(r.min_row, r.max_row, r.min_col, r.max_col)] = (r.min_row, r.max_row-1, r.min_col, r.max_col)
            elif r.min_row> row:
                map[(r.min_row, r.max_row, r.min_col, r.max_col)] = (r.min_row - 1, r.max_row - 1, r.min_col, r.max_col)

        for r in removelist:
            sheet.unmerge_cells(start_row=r.min_row, start_column=r.min_col, end_row=r.max_row, end_column=r.max_col)
        for min_row, max_row, min_col, max_col in map.keys():
            sheet.unmerge_cells(start_row= min_row, start_column=min_col, end_row=max_row, end_column= max_col)
        for min_row, max_row, min_col, max_col in map.values():
            sheet.merge_cells(start_row= min_row, start_column=min_col, end_row=max_row, end_column= max_col)

        for i in range(row, sheet.max_row+1):
            sheet.row_dimensions[i].height = sheet.row_dimensions[i+1].height if sheet.row_dimensions[i+1].height is not None else 15

    def mergeCell(self, startRow, endRow, startCol, endCol):
        sheet: Worksheet = self.worksheet
        sheet.merge_cells(start_row = startRow, end_row= endRow, start_column= startCol, end_column= endCol)

    def save(self, path):
        borderfix(self.workbook)
        self.workbook.save(path)

if __name__=='__main__':

    processor = OpenpyXlExcelProcessor('xlstemplates/template2.xlsx','Sheet1')
    evaluator = PyEvaluator()
    embeddedEvaluator = EmbeddedPyEvaluator(evaluator)
    context = DictExcelTemplateContext({
        'parts': [
            {'name': 'p1', 'items': [
                {'itemname': 'p11', 'v': 'sdf1'},
                {'itemname': 'p12', 'v': 'sdf2'},
                {'itemname': 'p13', 'v': 'sdf3'},
                {'itemname': 'p14', 'v': 'sdf3'},
            ]},
            {'name': 'p2', 'items': [
                {'itemname': 'p21', 'v': 'sdd1'},
            ]},
            {'name': 'p3', 'items': [
                {'itemname': 'p31', 'v': 'sdf1'},
                {'itemname': 'p32', 'v': 'sdf2'},
                {'itemname': 'p33', 'v': 'sdf3'},
            ]},

        ]
    })
    archtemplate = ExcelArchetectureTemplate(processor, context, evaluator)
    rptemplate = ExcelRelpaceTemplate(processor, context, embeddedEvaluator)
    archtemplate.process()
    rptemplate.process()
    processor.save('out5.xlsx')
    pass
